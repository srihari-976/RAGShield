"""Document text extraction with per-type loaders.

Supported: PDF (text + OCR fallback for scanned/image PDFs), DOCX, XLSX, XLS,
CSV, TXT, MD, HTML, JSON.
"""

import csv
import io
import json
from pathlib import Path

import pymupdf
from bs4 import BeautifulSoup
from docx import Document as DocxDocument
from pypdf import PdfReader

from app.core.config import get_settings

settings = get_settings()

ALLOWED_EXTENSIONS = {
    ".pdf", ".docx", ".xlsx", ".xls", ".txt", ".md", ".html", ".htm", ".json", ".csv",
}


def detect_mime_type(filename: str) -> str:
    ext = Path(filename).suffix.lower()
    return {
        ".pdf": "application/pdf",
        ".docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ".xls": "application/vnd.ms-excel",
        ".txt": "text/plain",
        ".md": "text/markdown",
        ".html": "text/html",
        ".htm": "text/html",
        ".json": "application/json",
        ".csv": "text/csv",
    }.get(ext, "application/octet-stream")


class ExtractionError(Exception):
    pass


def extract_text(filename: str, content: bytes) -> str:
    ext = Path(filename).suffix.lower()
    if ext == ".pdf":
        return _extract_pdf(content)
    if ext == ".docx":
        return _extract_docx(content)
    if ext == ".xlsx":
        return _extract_xlsx(content)
    if ext == ".xls":
        return _extract_xls(content)
    if ext in (".txt", ".md"):
        return content.decode("utf-8", errors="replace")
    if ext in (".html", ".htm"):
        soup = BeautifulSoup(content, "html.parser")
        for tag in soup(["script", "style"]):
            tag.decompose()
        return soup.get_text(separator="\n")
    if ext == ".json":
        try:
            return json.dumps(json.loads(content), indent=2, ensure_ascii=False)
        except json.JSONDecodeError as e:
            raise ExtractionError(f"invalid JSON: {e}") from e
    if ext == ".csv":
        return _extract_csv(content)
    raise ExtractionError(f"unsupported file type {ext}")


def _extract_pdf(content: bytes) -> str:
    """Text-first extraction; falls back to OCR for image/scanned PDFs."""
    try:
        reader = PdfReader(io.BytesIO(content))
        pages = [page.extract_text() or "" for page in reader.pages]
        text = "\n\n".join(pages).strip()
        if text:
            return text
    except Exception as e:
        raise ExtractionError(f"PDF extraction failed: {e}") from e

    # No embedded text layer — the PDF is scanned/image-based. OCR it.
    return _ocr_pdf(content)


def _ocr_pdf(content: bytes) -> str:
    try:
        from rapidocr_onnxruntime import RapidOCR
    except ImportError as e:  # pragma: no cover
        raise ExtractionError(
            "scanned PDF detected but OCR is unavailable: install rapidocr_onnxruntime"
        ) from e

    ocr = RapidOCR()
    parts: list[str] = []
    with pymupdf.open(stream=content, filetype="pdf") as doc:
        for page in doc:
            pix = page.get_pixmap(dpi=200)
            img_bytes = pix.tobytes("png")
            result, _ = ocr(img_bytes)
            if result:
                page_text = "\n".join(line[1] for line in result)
                parts.append(page_text)
    text = "\n\n".join(parts).strip()
    if not text:
        raise ExtractionError("no text could be extracted from the document (OCR returned empty)")
    return text


def _extract_docx(content: bytes) -> str:
    try:
        doc = DocxDocument(io.BytesIO(content))
        return "\n".join(p.text for p in doc.paragraphs if p.text)
    except Exception as e:
        raise ExtractionError(f"DOCX extraction failed: {e}") from e


def _extract_xlsx(content: bytes) -> str:
    try:
        from openpyxl import load_workbook
    except ImportError as e:  # pragma: no cover
        raise ExtractionError("XLSX support requires openpyxl") from e
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        lines: list[str] = []
        for sheet in wb.worksheets:
            lines.append(f"--- Sheet: {sheet.title} ---")
            for row in sheet.iter_rows(values_only=True):
                values = ["" if v is None else str(v) for v in row]
                if any(v.strip() for v in values):
                    lines.append("\t".join(values))
        return "\n".join(lines)
    except Exception as e:
        raise ExtractionError(f"XLSX extraction failed: {e}") from e


def _extract_xls(content: bytes) -> str:
    try:
        import xlrd
    except ImportError as e:  # pragma: no cover
        raise ExtractionError("XLS support requires xlrd") from e
    try:
        book = xlrd.open_workbook(file_contents=content)
        lines: list[str] = []
        for sheet in book.sheets():
            lines.append(f"--- Sheet: {sheet.name} ---")
            for row_idx in range(sheet.nrows):
                values = [str(cell.value) for cell in sheet.row(row_idx)]
                if any(v.strip() for v in values):
                    lines.append("\t".join(values))
        return "\n".join(lines)
    except Exception as e:
        raise ExtractionError(f"XLS extraction failed: {e}") from e


def _extract_csv(content: bytes) -> str:
    text = content.decode("utf-8", errors="replace")
    reader = csv.reader(io.StringIO(text))
    lines = [",".join(row) for row in reader]
    return "\n".join(lines)
